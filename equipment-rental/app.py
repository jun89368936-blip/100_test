from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, session
from datetime import datetime, date
from db import init_db, get_db, INTEGRITY_ERROR
from telegram_notify import notify_rent, notify_return
from werkzeug.security import generate_password_hash, check_password_hash
from dotenv import load_dotenv
import os
import threading
import time
import requests as _requests

load_dotenv()

app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", "rental-secret-key-change-in-prod")


def _keep_alive():
    """Render 무료 플랜 cold start 방지: 4분마다 자기 자신에게 핑"""
    time.sleep(60)  # 앱 완전 기동 대기
    base = os.getenv("RENDER_EXTERNAL_URL", "")
    if not base:
        return  # 로컬에서는 실행 안 함
    while True:
        try:
            _requests.get(f"{base}/health", timeout=10)
        except Exception:
            pass
        time.sleep(240)  # 4분 간격


threading.Thread(target=_keep_alive, daemon=True).start()
ADMIN_PIN = os.getenv("ADMIN_PIN", "1234")


def login_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("user_id"):
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated


def admin_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("is_admin"):
            return redirect(url_for("admin_login"))
        return f(*args, **kwargs)
    return decorated

ITEM_COLORS = ["#3788d8", "#e67e22", "#27ae60", "#8e44ad", "#c0392b"]


@app.template_filter("fmt_date")
def fmt_date(value):
    if not value:
        return "—"
    try:
        from datetime import datetime as dt
        d = dt.fromisoformat(str(value)[:10])
        return f"{d.year % 100:02d}년 {d.month:02d}월 {d.day:02d}일"
    except Exception:
        return value


@app.route("/health")
def health():
    return "ok", 200


@app.route("/login", methods=["GET", "POST"])
def login():
    if session.get("user_id"):
        return redirect(url_for("index"))
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()
        conn = get_db()
        user = conn.execute(
            "SELECT * FROM users WHERE username = ?", (username,)
        ).fetchone()
        conn.close()
        if user and check_password_hash(user["password_hash"], password):
            session["user_id"] = user["id"]
            session["display_name"] = user["display_name"]
            return redirect(url_for("index"))
        flash("아이디 또는 비밀번호가 올바르지 않습니다.", "error")
    return render_template("login.html")


@app.route("/logout")
def logout():
    session.pop("user_id", None)
    session.pop("display_name", None)
    return redirect(url_for("login"))


@app.route("/profile/change-password", methods=["GET", "POST"])
@login_required
def change_password():
    if request.method == "POST":
        current = request.form.get("current_password", "").strip()
        new_pw = request.form.get("new_password", "").strip()
        confirm = request.form.get("confirm_password", "").strip()
        if not current or not new_pw or not confirm:
            flash("모든 항목을 입력해주세요.", "error")
            return redirect(url_for("change_password"))
        if new_pw != confirm:
            flash("새 비밀번호가 일치하지 않습니다.", "error")
            return redirect(url_for("change_password"))
        if len(new_pw) < 4:
            flash("비밀번호는 4자리 이상이어야 합니다.", "error")
            return redirect(url_for("change_password"))
        conn = get_db()
        user = conn.execute("SELECT * FROM users WHERE id = ?", (session["user_id"],)).fetchone()
        if not user or not check_password_hash(user["password_hash"], current):
            conn.close()
            flash("현재 비밀번호가 올바르지 않습니다.", "error")
            return redirect(url_for("change_password"))
        conn.execute(
            "UPDATE users SET password_hash = ? WHERE id = ?",
            (generate_password_hash(new_pw), session["user_id"]),
        )
        conn.commit()
        conn.close()
        flash("비밀번호가 변경되었습니다.", "success")
        return redirect(url_for("index"))
    return render_template("change_password.html")


_db_initialized = False

@app.before_request
def startup():
    global _db_initialized
    if not _db_initialized:
        try:
            init_db()
        except Exception:
            import traceback, sys
            print("[init_db FAILED]", file=sys.stderr)
            traceback.print_exc()
        _db_initialized = True


@app.errorhandler(Exception)
def handle_exception(e):
    """예외 트레이스백을 로그로 남긴다 (Render 로그에서 확인 가능)."""
    from werkzeug.exceptions import HTTPException
    if isinstance(e, HTTPException):
        return e
    import traceback, sys
    print(f"[ERROR] {request.method} {request.path}", file=sys.stderr)
    traceback.print_exc()
    sys.stderr.flush()
    if "could not translate host name" in str(e) or "could not connect" in str(e):
        return (
            "데이터베이스에 연결할 수 없습니다.\n\n"
            "Render의 PostgreSQL 인스턴스가 만료·삭제되었을 수 있습니다.\n"
            "새 데이터베이스를 생성한 뒤 DATABASE_URL 환경변수를 교체해주세요.\n\n"
            f"({type(e).__name__})", 503,
            {"Content-Type": "text/plain; charset=utf-8"})
    return ("일시적인 오류가 발생했습니다. 잠시 후 다시 시도해주세요.", 500,
            {"Content-Type": "text/plain; charset=utf-8"})


@app.route("/health/db")
def health_db():
    """DB 연결·스키마 진단 (문제 해결 후 제거 예정)."""
    import traceback
    lines = []
    try:
        from db import _USE_PG
        lines.append(f"backend: {'postgresql' if _USE_PG else 'sqlite'}")
    except Exception as e:
        lines.append(f"backend: unknown ({e})")
    try:
        conn = get_db()
        lines.append("connect: OK")
        for table in ("users", "items", "rentals"):
            try:
                n = conn.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0]
                lines.append(f"{table}: {n} rows")
            except Exception as e:
                try:
                    conn._conn.rollback()
                except Exception:
                    pass
                lines.append(f"{table}: ERROR {type(e).__name__}: {e}")
        for table, col in (("items", "sort_order"), ("rentals", "cancelled_at")):
            try:
                conn.execute(f"SELECT {col} FROM {table} LIMIT 1").fetchone()
                lines.append(f"{table}.{col}: OK")
            except Exception as e:
                try:
                    conn._conn.rollback()
                except Exception:
                    pass
                lines.append(f"{table}.{col}: MISSING {type(e).__name__}: {e}")
        conn.close()
    except Exception:
        lines.append("connect: FAILED")
        lines.append(traceback.format_exc())
    return ("\n".join(lines), 200, {"Content-Type": "text/plain; charset=utf-8"})


@app.route("/")
@login_required
def index():
    from datetime import timedelta
    today = date.today().isoformat()
    display_name = session.get("display_name", "")
    conn = get_db()

    # 물품 목록 (정확히 N개, 중복 없음)
    items = conn.execute(
        "SELECT * FROM items ORDER BY sort_order, id"
    ).fetchall()

    # 현재 활성/예약 대여: item_id별 최초 1건 (가장 빠른 시작일)
    active_rows = conn.execute("""
        SELECT r.item_id, r.borrower_name, r.rented_at, r.due_date
        FROM rentals r
        WHERE r.returned_at IS NULL AND r.cancelled_at IS NULL
          AND (r.due_date IS NULL OR r.due_date >= ?)
        ORDER BY r.rented_at ASC
    """, (today,)).fetchall()
    rental_by_item = {}
    for r in active_rows:
        if r["item_id"] not in rental_by_item:
            rental_by_item[r["item_id"]] = r

    my_rentals = conn.execute("""
        SELECT r.id, r.item_id, i.name AS item_name, i.type AS item_type,
               r.rented_at, r.due_date
        FROM rentals r JOIN items i ON i.id = r.item_id
        WHERE r.borrower_name = ? AND r.returned_at IS NULL AND r.cancelled_at IS NULL
        ORDER BY r.due_date ASC
    """, (display_name,)).fetchall()

    # 주간 대여현황: 이번 주 월~일 (7일)
    weekday = date.today().weekday()
    week_start = date.today() - timedelta(days=weekday)
    week_days = [week_start + timedelta(days=i) for i in range(7)]
    week_labels = ["월", "화", "수", "목", "금", "토", "일"]
    week_info = list(zip([d.isoformat() for d in week_days],
                         [d.strftime('%m/%d') for d in week_days],
                         week_labels))

    week_rentals = conn.execute("""
        SELECT r.item_id, r.borrower_name, r.rented_at, r.due_date
        FROM rentals r
        WHERE r.returned_at IS NULL AND r.cancelled_at IS NULL
          AND r.rented_at <= ? AND (r.due_date IS NULL OR r.due_date >= ?)
    """, (week_days[-1].isoformat(), week_days[0].isoformat())).fetchall()

    # {item_id: {date_str: borrower_name}}
    week_map = {}
    for r in week_rentals:
        try:
            s = date.fromisoformat(str(r["rented_at"])[:10])
            e = date.fromisoformat(str(r["due_date"])[:10]) if r["due_date"] else date.today()
        except Exception:
            continue
        for d in week_days:
            if s <= d <= e:
                week_map.setdefault(r["item_id"], {})[d.isoformat()] = r["borrower_name"]

    conn.close()
    max_date = (date.today() + timedelta(days=31)).isoformat()
    return render_template("index.html", items=items, today=today, my_rentals=my_rentals,
                           max_date=max_date, display_name=display_name,
                           rental_by_item=rental_by_item,
                           week_info=week_info, week_map=week_map)


@app.route("/calendar")
@login_required
def calendar_view():
    conn = get_db()
    items = conn.execute("SELECT * FROM items ORDER BY id").fetchall()
    conn.close()
    return render_template("calendar.html", items=items, colors=ITEM_COLORS)


@app.route("/admin/calendar")
@admin_required
def admin_calendar():
    conn = get_db()
    items = conn.execute("SELECT * FROM items ORDER BY sort_order, id").fetchall()
    conn.close()
    return render_template("admin/calendar.html", items=items, colors=ITEM_COLORS)


@app.route("/api/events")
@login_required
def api_events():
    conn = get_db()
    rows = conn.execute("""
        SELECT r.id, r.item_id, r.borrower_name, r.rented_at, r.returned_at, r.due_date,
               r.cancelled_at, i.name AS item_name
        FROM rentals r
        JOIN items i ON i.id = r.item_id
        WHERE r.cancelled_at IS NULL
        ORDER BY r.id DESC
    """).fetchall()
    conn.close()

    events = []
    for idx, r in enumerate(rows):
        start = r["rented_at"][:10] if r["rented_at"] else None
        # FullCalendar end is exclusive — add 1 day so due_date shows fully
        from datetime import timedelta as _td
        end = None
        if r["returned_at"]:
            try:
                end = (date.fromisoformat(r["returned_at"][:10]) + _td(days=1)).isoformat()
            except Exception:
                end = r["returned_at"][:10]
        elif r["due_date"]:
            try:
                end = (date.fromisoformat(r["due_date"][:10]) + _td(days=1)).isoformat()
            except Exception:
                end = r["due_date"]
        if not end:
            end = (date.today() + _td(days=1)).isoformat()

        is_returned = bool(r["returned_at"])
        color = "#b0b8c8" if is_returned else ITEM_COLORS[(r["item_id"] - 1) % len(ITEM_COLORS)]
        status = "반납완료" if is_returned else "대여중"
        events.append({
            "id": r["id"],
            "title": f"{'[반납] ' if is_returned else ''}{r['item_name']} ({r['borrower_name']})",
            "start": start,
            "end": end,
            "color": color,
            "textColor": "#fff" if not is_returned else "#6b7280",
            "extendedProps": {"status": status, "borrower": r["borrower_name"]},
        })

    return jsonify(events)


@app.route("/rent/<int:item_id>", methods=["POST"])
@login_required
def rent(item_id):
    borrower_name = session.get("display_name", "")
    rented_at = request.form.get("rented_at", "").strip()
    due_date = request.form.get("due_date", "").strip()
    if not borrower_name:
        flash("로그인이 필요합니다.", "error")
        return redirect(url_for("login"))
    if not rented_at:
        flash("대여일을 선택해주세요.", "error")
        return redirect(url_for("index"))
    if not due_date:
        flash("반납일을 선택해주세요.", "error")
        return redirect(url_for("index"))

    conn = get_db()
    # 본인이 이미 신청(대여/예약)한 동일 물품 중복 신청 차단
    duplicate = conn.execute("""
        SELECT COUNT(*) FROM rentals
        WHERE item_id = ?
          AND borrower_name = ?
          AND returned_at IS NULL
          AND cancelled_at IS NULL
    """, (item_id, borrower_name)).fetchone()[0]
    if duplicate:
        flash("이미 신청한 물품입니다 — 같은 물품은 중복 신청할 수 없습니다.", "error")
        conn.close()
        return redirect(url_for("index"))
    # 날짜 범위 겹침 체크 (반납예정일 당일은 새 대여 허용)
    conflict = conn.execute("""
        SELECT COUNT(*) FROM rentals
        WHERE item_id = ?
          AND returned_at IS NULL
          AND cancelled_at IS NULL
          AND rented_at < ?
          AND (due_date IS NULL OR due_date > ?)
    """, (item_id, due_date, rented_at)).fetchone()[0]
    if conflict:
        flash("예약불가 — 해당 기간에 이미 예약된 물품입니다.", "error")
        conn.close()
        return redirect(url_for("index"))
    try:
        conn.execute(
            "INSERT INTO rentals (item_id, borrower_name, rented_at, due_date) VALUES (?, ?, ?, ?)",
            (item_id, borrower_name, rented_at, due_date),
        )
        conn.commit()
        item = conn.execute("SELECT name FROM items WHERE id = ?", (item_id,)).fetchone()
        flash("대여가 완료되었습니다.", "success")
        notify_rent(item["name"], borrower_name, rented_at)
    except Exception:
        flash("대여 처리 중 오류가 발생했습니다.", "error")
    finally:
        conn.close()

    return redirect(url_for("index"))


@app.route("/return/<int:item_id>", methods=["POST"])
@login_required
def return_item(item_id):
    conn = get_db()
    returned_at = datetime.now().isoformat(timespec="seconds")
    rental = conn.execute(
        "SELECT r.borrower_name, i.name FROM rentals r JOIN items i ON i.id = r.item_id "
        "WHERE r.item_id = ? AND r.returned_at IS NULL AND r.cancelled_at IS NULL",
        (item_id,),
    ).fetchone()
    conn.execute(
        "UPDATE rentals SET returned_at = ? WHERE item_id = ? AND returned_at IS NULL AND cancelled_at IS NULL",
        (returned_at, item_id),
    )
    conn.commit()
    conn.close()
    flash("반납이 완료되었습니다.")
    if rental:
        notify_return(rental["name"], rental["borrower_name"], returned_at)
    return redirect(url_for("index"))


@app.route("/cancel/<int:rental_id>", methods=["POST"])
@login_required
def cancel_rental(rental_id):
    conn = get_db()
    rental = conn.execute(
        "SELECT r.id, r.borrower_name, r.returned_at, r.cancelled_at, i.name AS item_name "
        "FROM rentals r JOIN items i ON i.id = r.item_id WHERE r.id = ?",
        (rental_id,),
    ).fetchone()
    if not rental:
        flash("존재하지 않는 대여 기록입니다.", "error")
        conn.close()
        return redirect(url_for("index"))
    if rental["returned_at"] is not None:
        flash("이미 반납된 대여입니다.", "error")
        conn.close()
        return redirect(url_for("index"))
    if rental.get("cancelled_at") is not None:
        flash("이미 취소된 대여입니다.", "error")
        conn.close()
        return redirect(url_for("index"))
    if rental["borrower_name"] != session.get("display_name"):
        flash("본인이 신청한 대여만 취소할 수 있습니다.", "error")
        conn.close()
        return redirect(url_for("index"))
    conn.execute(
        "UPDATE rentals SET cancelled_at = ? WHERE id = ?",
        (datetime.now().isoformat(timespec="seconds"), rental_id),
    )
    conn.commit()
    conn.close()
    flash(f"'{rental['item_name']}' 대여 신청이 취소되었습니다.", "success")
    return redirect(url_for("index"))


@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        pin = request.form.get("pin", "").strip()
        if pin == ADMIN_PIN:
            session["is_admin"] = True
            return redirect(url_for("admin_dashboard"))
        flash("PIN이 올바르지 않습니다.")
    return render_template("admin_login.html")


@app.route("/admin/logout")
def admin_logout():
    session.pop("is_admin", None)
    return redirect(url_for("index"))


def _calc_monthly_rental_days(conn, items):
    """물품별 월별 대여일수 계산. {item_name: {YYYY-MM: days}}"""
    from datetime import datetime as dt, timedelta
    rows = conn.execute("""
        SELECT r.item_id, r.rented_at, r.due_date, r.returned_at
        FROM rentals r
        WHERE r.cancelled_at IS NULL
        ORDER BY r.item_id, r.rented_at
    """).fetchall()

    today = date.today()
    data = {item["name"]: {} for item in items}
    id_to_name = {item["id"]: item["name"] for item in items}

    for r in rows:
        name = id_to_name.get(r["item_id"])
        if not name:
            continue
        try:
            start = dt.fromisoformat(str(r["rented_at"])[:10]).date()
        except Exception:
            continue
        if r["returned_at"]:
            try:
                end = dt.fromisoformat(str(r["returned_at"])[:10]).date()
            except Exception:
                end = today
        elif r["due_date"]:
            try:
                # 아직 반납 안 된 건은 오늘까지만 집계 (미래 일수 제외)
                end = min(dt.fromisoformat(str(r["due_date"])[:10]).date(), today)
            except Exception:
                end = today
        else:
            end = today

        cur = start
        while cur <= end:
            key = cur.strftime("%Y-%m")
            data[name][key] = data[name].get(key, 0) + 1
            cur += timedelta(days=1)

    return data


@app.route("/admin")
@admin_required
def admin_dashboard():
    today_str = date.today().isoformat()
    conn = get_db()
    items = conn.execute("SELECT id, name FROM items ORDER BY sort_order, id").fetchall()
    total = len(items)
    rented_today = conn.execute("""
        SELECT COUNT(DISTINCT r.item_id) FROM rentals r
        WHERE r.returned_at IS NULL AND r.cancelled_at IS NULL
          AND r.rented_at <= ? AND (r.due_date IS NULL OR r.due_date >= ?)
    """, (today_str, today_str)).fetchone()[0]
    total_rentals = conn.execute(
        "SELECT COUNT(*) FROM rentals WHERE cancelled_at IS NULL"
    ).fetchone()[0]
    active_rentals = conn.execute("""
        SELECT r.id AS rental_id, r.borrower_name, r.rented_at, r.due_date, i.name AS item_name
        FROM rentals r JOIN items i ON i.id = r.item_id
        WHERE r.returned_at IS NULL AND r.cancelled_at IS NULL
        ORDER BY r.rented_at DESC
    """).fetchall()

    monthly_data = _calc_monthly_rental_days(conn, items)
    conn.close()

    # 월 목록 (최근 6개월)
    months = []
    for i in range(5, -1, -1):
        d = date.today().replace(day=1)
        m = d.month - i
        y = d.year
        while m <= 0:
            m += 12
            y -= 1
        months.append(f"{y}-{m:02d}")

    # 월별 합계 미리 계산 (Jinja2 namespace 미사용)
    monthly_totals = {}
    for m in months:
        monthly_totals[m] = sum(
            monthly_data.get(item["name"], {}).get(m, 0) for item in items
        )

    stats = {"total": total, "rented": rented_today,
             "available": total - rented_today, "total_rentals": total_rentals}
    return render_template("admin/dashboard.html", stats=stats, active_rentals=active_rentals,
                           today=today_str, items=items,
                           monthly_data=monthly_data, months=months,
                           monthly_totals=monthly_totals)


@app.route("/admin/dashboard/export")
@admin_required
def admin_dashboard_export():
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from flask import send_file
    from datetime import timedelta, datetime as dt

    conn = get_db()
    items = conn.execute("SELECT id, name FROM items ORDER BY sort_order, id").fetchall()
    monthly_data = _calc_monthly_rental_days(conn, items)
    conn.close()

    # 최근 12개월
    months = []
    for i in range(11, -1, -1):
        d = date.today().replace(day=1)
        m = d.month - i
        y = d.year
        while m <= 0:
            m += 12
            y -= 1
        months.append(f"{y}-{m:02d}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "월별 대여일수"

    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="1E293B")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    sub_fill = PatternFill("solid", fgColor="3B4A6B")
    alt_fill = PatternFill("solid", fgColor="F8FAFC")

    # 타이틀
    ws.merge_cells(f"A1:{get_column_letter(len(months)+1)}1")
    tc = ws["A1"]
    tc.value = "물품별 월별 대여일수"
    tc.font = Font(bold=True, size=14, color="1E293B")
    tc.alignment = center
    ws.row_dimensions[1].height = 30

    # 헤더행: 물품명 | 월1 | 월2 | ...
    ws.cell(row=2, column=1, value="물품명").fill = hdr_fill
    ws.cell(row=2, column=1).font = hdr_font
    ws.cell(row=2, column=1).alignment = center
    ws.cell(row=2, column=1).border = border
    ws.column_dimensions["A"].width = 16
    for ci, m in enumerate(months, 2):
        cell = ws.cell(row=2, column=ci, value=m)
        cell.fill = sub_fill
        cell.font = hdr_font
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(ci)].width = 11
    ws.row_dimensions[2].height = 20

    # 데이터
    for ri, item in enumerate(items):
        row_num = ri + 3
        ws.cell(row=row_num, column=1, value=item["name"]).font = Font(bold=True)
        ws.cell(row=row_num, column=1).border = border
        ws.cell(row=row_num, column=1).alignment = center
        if ri % 2 == 1:
            ws.cell(row=row_num, column=1).fill = alt_fill
        for ci, m in enumerate(months, 2):
            days = monthly_data.get(item["name"], {}).get(m, 0)
            cell = ws.cell(row=row_num, column=ci, value=days if days else "")
            cell.alignment = center
            cell.border = border
            if ri % 2 == 1:
                cell.fill = alt_fill
            if days:
                cell.font = Font(color="1E40AF", bold=True)
        ws.row_dimensions[row_num].height = 20

    # 합계행
    sum_row = len(items) + 3
    ws.cell(row=sum_row, column=1, value="합계").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=sum_row, column=1).fill = PatternFill("solid", fgColor="7C3AED")
    ws.cell(row=sum_row, column=1).alignment = center
    ws.cell(row=sum_row, column=1).border = border
    for ci, m in enumerate(months, 2):
        total_days = sum(monthly_data.get(item["name"], {}).get(m, 0) for item in items)
        cell = ws.cell(row=sum_row, column=ci, value=total_days if total_days else "")
        cell.fill = PatternFill("solid", fgColor="EDE9FE")
        cell.font = Font(bold=True, color="4C1D95")
        cell.alignment = center
        cell.border = border

    ws.cell(row=sum_row + 2, column=1,
            value=f"출력일: {date.today().strftime('%Y년 %m월 %d일')}").font = Font(color="94A3B8", size=9)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"월별대여일수_{date.today().strftime('%Y%m%d')}.xlsx"
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=filename)


@app.route("/admin/items")
@admin_required
def admin_items():
    conn = get_db()
    today = date.today().isoformat()
    items = conn.execute("""
        SELECT i.*,
               (SELECT COUNT(*) FROM rentals r
                WHERE r.item_id = i.id AND r.returned_at IS NULL
                  AND r.rented_at <= ? AND (r.due_date IS NULL OR r.due_date > ?)) AS is_rented
        FROM items i ORDER BY i.sort_order, i.id
    """, (today, today)).fetchall()
    conn.close()
    return render_template("admin/items.html", items=items)


@app.route("/admin/items/add", methods=["POST"])
@admin_required
def admin_add_item():
    name = request.form.get("name", "").strip()
    item_type = request.form.get("type", "").strip()
    if not name or not item_type:
        flash("물품명과 종류를 모두 입력해주세요.", "error")
        return redirect(url_for("admin_items"))
    conn = get_db()
    max_order = conn.execute("SELECT COALESCE(MAX(sort_order), -1) FROM items").fetchone()[0]
    conn.execute("INSERT INTO items (name, type, sort_order) VALUES (?, ?, ?)", (name, item_type, max_order + 1))
    conn.commit()
    conn.close()
    flash(f"'{name}' 물품이 추가되었습니다.", "success")
    return redirect(url_for("admin_items"))


@app.route("/admin/items/move/<int:item_id>/<direction>", methods=["POST"])
@admin_required
def admin_move_item(item_id, direction):
    conn = get_db()
    items = conn.execute("SELECT id, sort_order FROM items ORDER BY sort_order, id").fetchall()
    ids = [r["id"] for r in items]
    if item_id not in ids:
        conn.close()
        return redirect(url_for("admin_items"))
    idx = ids.index(item_id)
    swap_idx = idx - 1 if direction == "up" else idx + 1
    if 0 <= swap_idx < len(ids):
        a, b = items[idx], items[swap_idx]
        conn.execute("UPDATE items SET sort_order = ? WHERE id = ?", (b["sort_order"], a["id"]))
        conn.execute("UPDATE items SET sort_order = ? WHERE id = ?", (a["sort_order"], b["id"]))
        conn.commit()
    conn.close()
    return redirect(url_for("admin_items"))


@app.route("/admin/items/delete/<int:item_id>", methods=["POST"])
@admin_required
def admin_delete_item(item_id):
    conn = get_db()
    is_rented = conn.execute(
        "SELECT COUNT(*) FROM rentals WHERE item_id = ? AND returned_at IS NULL", (item_id,)
    ).fetchone()[0]
    if is_rented:
        flash("현재 대여 중인 물품은 삭제할 수 없습니다.", "error")
        conn.close()
        return redirect(url_for("admin_items"))
    item = conn.execute("SELECT name FROM items WHERE id = ?", (item_id,)).fetchone()
    conn.execute("DELETE FROM rentals WHERE item_id = ?", (item_id,))
    conn.execute("DELETE FROM items WHERE id = ?", (item_id,))
    conn.commit()
    conn.close()
    flash(f"'{item['name']}' 물품이 삭제되었습니다.", "success")
    return redirect(url_for("admin_items"))


@app.route("/admin/rentals")
@admin_required
def admin_rentals():
    item_filter = request.args.get("item_id", "")
    conn = get_db()
    if item_filter:
        rows = conn.execute("""
            SELECT r.id, i.name AS item_name, i.type AS item_type,
                   r.borrower_name, r.rented_at, r.returned_at, r.due_date, r.cancelled_at
            FROM rentals r JOIN items i ON i.id = r.item_id
            WHERE r.item_id = ?
            ORDER BY r.id DESC
        """, (item_filter,)).fetchall()
    else:
        rows = conn.execute("""
            SELECT r.id, i.name AS item_name, i.type AS item_type,
                   r.borrower_name, r.rented_at, r.returned_at, r.due_date, r.cancelled_at
            FROM rentals r JOIN items i ON i.id = r.item_id
            ORDER BY r.id DESC
        """).fetchall()
    available_items = conn.execute("""
        SELECT i.id, i.name FROM items i
        WHERE NOT EXISTS (
            SELECT 1 FROM rentals r WHERE r.item_id = i.id AND r.returned_at IS NULL
        )
        ORDER BY i.id
    """).fetchall()
    all_items = conn.execute("SELECT id, name FROM items ORDER BY sort_order, id").fetchall()
    conn.close()
    return render_template("admin/rentals.html", rows=rows, available_items=available_items,
                           all_items=all_items, item_filter=item_filter,
                           today=date.today().isoformat())


@app.route("/admin/rentals/add", methods=["POST"])
@admin_required
def admin_add_rental():
    item_id = request.form.get("item_id", "").strip()
    borrower_name = request.form.get("borrower_name", "").strip()
    rented_at = request.form.get("rented_at", "").strip() or date.today().isoformat()
    due_date = request.form.get("due_date", "").strip()
    if not item_id or not borrower_name:
        flash("물품과 대여자 이름을 입력해주세요.", "error")
        return redirect(url_for("admin_rentals"))
    conn = get_db()
    conflict = conn.execute("""
        SELECT COUNT(*) FROM rentals
        WHERE item_id = ?
          AND returned_at IS NULL
          AND rented_at < ?
          AND (due_date IS NULL OR due_date > ?)
    """, (item_id, due_date or "9999-12-31", rented_at)).fetchone()[0]
    if conflict:
        flash("해당 기간에 이미 대여 중인 물품입니다.", "error")
        conn.close()
        return redirect(url_for("admin_rentals"))
    try:
        conn.execute(
            "INSERT INTO rentals (item_id, borrower_name, rented_at, due_date) VALUES (?, ?, ?, ?)",
            (item_id, borrower_name, rented_at, due_date or None),
        )
        conn.commit()
        item = conn.execute("SELECT name FROM items WHERE id = ?", (item_id,)).fetchone()
        flash(f"'{item['name']}' 대여가 등록되었습니다.", "success")
        notify_rent(item["name"], borrower_name, rented_at)
    except Exception:
        flash("대여 처리 중 오류가 발생했습니다.", "error")
    finally:
        conn.close()
    return redirect(url_for("admin_rentals"))


@app.route("/admin/rentals/return/<int:rental_id>", methods=["POST"])
@admin_required
def admin_return_rental(rental_id):
    conn = get_db()
    returned_at = datetime.now().isoformat(timespec="seconds")
    rental = conn.execute(
        "SELECT r.borrower_name, i.name FROM rentals r JOIN items i ON i.id = r.item_id WHERE r.id = ?",
        (rental_id,),
    ).fetchone()
    conn.execute(
        "UPDATE rentals SET returned_at = ? WHERE id = ? AND returned_at IS NULL",
        (returned_at, rental_id),
    )
    conn.commit()
    conn.close()
    if rental:
        flash(f"'{rental['name']}' 반납 처리가 완료되었습니다.", "success")
        notify_return(rental["name"], rental["borrower_name"], returned_at)
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/rentals/delete/<int:rental_id>", methods=["POST"])
@admin_required
def admin_delete_rental(rental_id):
    conn = get_db()
    conn.execute("DELETE FROM rentals WHERE id = ?", (rental_id,))
    conn.commit()
    conn.close()
    flash("대여 기록이 삭제되었습니다.", "success")
    return redirect(url_for("admin_rentals"))


@app.route("/admin/rentals/delete-selected", methods=["POST"])
@admin_required
def admin_delete_selected():
    ids = request.form.getlist("selected_ids")
    if not ids:
        flash("선택된 항목이 없습니다.", "error")
        return redirect(url_for("admin_rentals"))
    conn = get_db()
    for rid in ids:
        conn.execute("DELETE FROM rentals WHERE id = ?", (rid,))
    conn.commit()
    conn.close()
    flash(f"{len(ids)}건의 기록이 삭제되었습니다.", "success")
    return redirect(url_for("admin_rentals"))


@app.route("/admin/rentals/export")
@admin_required
def admin_rentals_export():
    import io
    from flask import send_file
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    item_filter = request.args.get("item_id", "")
    conn = get_db()
    if item_filter:
        rows = conn.execute("""
            SELECT r.id, i.name AS item_name, i.type AS item_type,
                   r.borrower_name, r.rented_at, r.due_date, r.returned_at
            FROM rentals r JOIN items i ON i.id = r.item_id
            WHERE r.item_id = ?
            ORDER BY r.id DESC
        """, (item_filter,)).fetchall()
        item_name = conn.execute("SELECT name FROM items WHERE id = ?", (item_filter,)).fetchone()
        sheet_title = item_name["name"] if item_name else "물품별"
    else:
        rows = conn.execute("""
            SELECT r.id, i.name AS item_name, i.type AS item_type,
                   r.borrower_name, r.rented_at, r.due_date, r.returned_at
            FROM rentals r JOIN items i ON i.id = r.item_id
            ORDER BY r.id DESC
        """).fetchall()
        sheet_title = "전체"
    conn.close()

    TYPE_LABELS = {"drone": "드론", "laptop": "노트북", "camera": "카메라", "etc": "기타"}

    def fmt(val):
        if not val:
            return "—"
        try:
            from datetime import datetime as dt
            d = dt.fromisoformat(str(val)[:10])
            return f"{d.year % 100:02d}년 {d.month:02d}월 {d.day:02d}일"
        except Exception:
            return val

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]

    header_fill = PatternFill("solid", fgColor="1E293B")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["No.", "물품명", "종류", "대여자", "대여일", "반납 예정일", "상태"]
    col_widths = [6, 16, 10, 12, 16, 16, 10]

    doc_title = f"부서 공유 물품 대여 이력 — {sheet_title}"
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = doc_title
    title_cell.font = Font(bold=True, size=14, color="1E293B")
    title_cell.alignment = center
    ws.row_dimensions[1].height = 32

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[2].height = 22

    alt_fill = PatternFill("solid", fgColor="F8FAFC")
    for idx, r in enumerate(rows, 1):
        row_num = idx + 2
        status = "반납완료" if r["returned_at"] else "대여 중"
        values = [idx, r["item_name"], TYPE_LABELS.get(r["item_type"], "기타"),
                  r["borrower_name"], fmt(r["rented_at"]), fmt(r["due_date"]), status]
        fill = alt_fill if idx % 2 == 0 else None
        status_color = "27AE60" if r["returned_at"] else "E67E22"
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.alignment = center
            cell.border = border
            if fill:
                cell.fill = fill
            if col == 7:
                cell.font = Font(bold=True, color=status_color)
        ws.row_dimensions[row_num].height = 20

    from datetime import date as _date
    ws.cell(row=len(rows) + 4, column=1,
            value=f"출력일: {_date.today().strftime('%Y년 %m월 %d일')}").font = Font(color="94A3B8", size=9)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe = sheet_title.replace(" ", "_")
    filename = f"대여이력_{safe}_{_date.today().strftime('%Y%m%d')}.xlsx"
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=filename)


@app.route("/admin/rentals/delete-all", methods=["POST"])
@admin_required
def admin_delete_all_rentals():
    conn = get_db()
    count = conn.execute("SELECT COUNT(*) FROM rentals").fetchone()[0]
    conn.execute("DELETE FROM rentals")
    conn.commit()
    conn.close()
    flash(f"전체 {count}건의 대여 기록이 삭제되었습니다.", "success")
    return redirect(url_for("admin_rentals"))


try:
    from members_hashed import MEMBERS_HASHED as _MEMBERS_HASHED
except ImportError:
    _MEMBERS_HASHED = None

_MEMBERS = [['203108', '류재남'], ['221247', '정용선'], ['221313', '강정운'], ['224267', '반용수'], ['120008', '허길헌'], ['202090', '신정훈'], ['203009', '박광용'], ['203044', '김민희'], ['205044', '이병철'], ['210056', '심혜진'], ['213068', '조은해'], ['215134', '신성환'], ['216010', '윤민석'], ['219010', '임희상'], ['219116', '김원준'], ['220225', '안희윤'], ['221011', '양희승'], ['222098', '강민성'], ['224117', '심경환'], ['224118', '이다영'], ['225187', '김예나'], ['225292', '김소연'], ['226179', '김태문'], ['226180', '황선우'], ['195230', '김상훈'], ['195143', '이효섭'], ['195262', '최현기'], ['197011', '황윤희'], ['201059', '최정민'], ['204035', '이진우'], ['204143', '윤형일'], ['206061', '신동림'], ['206069', '안영민'], ['206100', '유태윤'], ['206135', '김광연'], ['206167', '신태호'], ['210107', '이미경'], ['214047', '장우준'], ['214059', '안현민'], ['217049', '이재호'], ['219001', '윤은비'], ['219088', '최재경'], ['220011', '조규연'], ['221060', '배연지'], ['221225', '김한주'], ['222074', '심주영'], ['223010', '정수빈'], ['223116', '황호재'], ['226175', '이종민'], ['197008', '장근영'], ['198008', '김두수'], ['201003', '강승현'], ['201088', '김홍기'], ['201090', '김성현'], ['203117', '박도영'], ['204007', '김택수'], ['204052', '김준한'], ['204073', '황충진'], ['205057', '강상진'], ['205074', '권도엽'], ['206005', '한현준'], ['206056', '박효정'], ['207004', '장동진'], ['207104', '우성훈'], ['207133', '조남흥'], ['208001', '이현석'], ['208123', '이기석'], ['211009', '박지용'], ['213011', '이종헌'], ['214125', '장홍규'], ['215001', '곽상호'], ['215003', '김재윤'], ['216006', '최현준'], ['216009', '민종원'], ['216134', '양우근'], ['217006', '이태용'], ['217009', '김범수'], ['217095', '조대환'], ['218005', '고영훈'], ['218200', '남기윤'], ['219007', '신창희'], ['219009', '명노준'], ['219085', '정현우'], ['219129', '김태경'], ['220004', '양민석'], ['220014', '문호건'], ['220167', '이영진'], ['220272', '변현종'], ['221008', '이용준'], ['221088', '박수완'], ['221192', '김동욱'], ['221193', 'Kavuma Andrew'], ['222002', '윤도영'], ['222004', '김수연'], ['222012', '하지철'], ['223001', '강경진'], ['223002', '유수민'], ['223005', '장인영'], ['224164', '강지수'], ['225001', '조민영'], ['225002', '홍건의'], ['226002', '송광현'], ['226004', '이준영'], ['195122', '김재웅'], ['195164', '백두현'], ['198009', '김옥자'], ['199074', '김호용'], ['199076', '박지훈'], ['201075', '오성환'], ['201089', '심성칠'], ['203006', '김소연'], ['204006', '박성준'], ['204172', '백종수'], ['205073', '김진중'], ['206004', '오창석'], ['206143', '허외탁'], ['207005', '정일환'], ['207102', '전원준'], ['207203', '손정훈'], ['208003', '박희준'], ['208125', '강상철'], ['209001', '손준성'], ['209002', '박성식'], ['210004', '홍기웅'], ['215002', '김원우'], ['215009', '이승용'], ['215170', '박인병'], ['216001', '이승진'], ['216004', '박영서'], ['217147', '김숙현'], ['218003', '김규연'], ['219002', '신정훈'], ['219004', '김혁준'], ['219005', '이태연'], ['219173', '전중현'], ['219202', '신홍철'], ['220002', '강진영'], ['220008', '신동수'], ['220233', '이명준'], ['220321', '이치훈'], ['221001', '김규리'], ['222003', '이해선'], ['222006', '안송희'], ['222009', '오민진'], ['222329', '오현석'], ['223003', '김대웅'], ['223004', '김주형'], ['223007', '전우현'], ['224004', '심명진'], ['225003', '곽예진'], ['225004', '박민준'], ['225005', '하민성'], ['225006', '김민지'], ['226003', '유이지'], ['226005', '이희재'], ['226007', '윤홍기'], ['196139', '임차현'], ['203195', '고재갑'], ['205012', '주현'], ['206159', '최성훈'], ['207001', '김영상'], ['207101', '문정현'], ['207209', '권기욱'], ['208171', '강상규'], ['208180', '김준호'], ['210154', '송대원'], ['213006', '안성재'], ['215121', '김은솔'], ['215136', '정지훈'], ['215179', '이상황'], ['215191', '노유진'], ['217007', '김태훈'], ['217008', '김영준'], ['218001', '정원제'], ['218004', '김대권'], ['218006', '김영범'], ['219003', '조진희'], ['219168', '조용구'], ['219240', '장영욱'], ['220001', '이유리'], ['220005', '정영현'], ['220006', '인채진'], ['220012', '정하정'], ['220258', '이재혁'], ['221002', '엄재훈'], ['222010', '박현우'], ['222011', '위다영'], ['222292', '송민재'], ['223008', '노동휘'], ['223009', '서동민'], ['223165', '조시훈'], ['224006', '정민수'], ['224007', '지예주'], ['224008', '최현우'], ['224268', '강의남'], ['224349', '임익태'], ['225007', '손승범'], ['225008', '이선렬'], ['225009', '최병엽'], ['226008', '강신영'], ['194144', '이준희'], ['195024', '이훈'], ['197177', '안혜원'], ['201043', '윤호순'], ['202074', '박재준'], ['203003', '양치열'], ['203113', '최영삼'], ['203114', '이상기'], ['203175', '이용철'], ['203180', '윤성호'], ['204036', '강상욱'], ['204054', '박재구'], ['204057', '신범규'], ['204058', '이주철'], ['204127', '최종영'], ['205011', '최원혁'], ['206001', '손창호'], ['207030', '박승노'], ['207157', '신영일'], ['208002', '김현봉'], ['208004', '위승경'], ['208005', '허환'], ['208006', '윤병우'], ['208172', '형정우'], ['214033', '김현진'], ['215008', '이동현'], ['215189', '권대희'], ['217003', '정준원'], ['218168', '심보연'], ['219008', '강수은'], ['219077', '황우진'], ['219107', '민철화'], ['219228', '김동춘'], ['220003', '윤호준'], ['220007', '임창섭'], ['220009', '김태호'], ['220010', '김성규'], ['220100', '김기진'], ['220174', '진경범'], ['220318', '황현진'], ['221006', '마종찬'], ['221010', '이상혁'], ['222001', '김동휘'], ['222007', '권재영'], ['222008', '배진성'], ['223006', '손정민'], ['224001', '박형호'], ['224003', '신동주'], ['224005', '이광현'], ['224275', '이영선'], ['226001', '김수아'], ['226006', '백창연'], ['198069', '서근득']]


@app.route("/admin/users/seed", methods=["POST"])
@admin_required
def admin_seed_users():
    try:
        # 미리 계산된 해시 사용 (타임아웃 방지)
        members = _MEMBERS_HASHED if _MEMBERS_HASHED else [
            (s, n, generate_password_hash(s)) for s, n in _MEMBERS
        ]
        conn = get_db()
        added = skipped = 0
        for sabun, name, pw_hash in members:
            if conn.execute("SELECT id FROM users WHERE username = ?", (sabun,)).fetchone():
                skipped += 1
                continue
            conn.execute(
                "INSERT INTO users (username, password_hash, display_name) VALUES (?, ?, ?)",
                (sabun, pw_hash, name),
            )
            added += 1
        conn.commit()
        conn.close()
        flash(f"부서원 등록 완료: {added}명 추가, {skipped}명 중복 건너뜀", "success")
    except Exception as e:
        flash(f"오류 발생: {e}", "error")
    return redirect(url_for("admin_users"))


@app.route("/admin/users")
@admin_required
def admin_users():
    conn = get_db()
    users = conn.execute("SELECT id, username, display_name FROM users ORDER BY id").fetchall()
    conn.close()
    return render_template("admin/users.html", users=users)


@app.route("/admin/users/add", methods=["POST"])
@admin_required
def admin_add_user():
    username = request.form.get("username", "").strip()
    display_name = request.form.get("display_name", "").strip()
    password = request.form.get("password", "").strip()
    if not username or not display_name or not password:
        flash("모든 항목을 입력해주세요.", "error")
        return redirect(url_for("admin_users"))
    pw_hash = generate_password_hash(password)
    conn = get_db()
    try:
        conn.execute(
            "INSERT INTO users (username, password_hash, display_name) VALUES (?, ?, ?)",
            (username, pw_hash, display_name),
        )
        conn.commit()
        flash(f"'{display_name}' 계정이 추가되었습니다.", "success")
    except INTEGRITY_ERROR:
        flash(f"이미 존재하는 아이디입니다: {username}", "error")
    finally:
        conn.close()
    return redirect(url_for("admin_users"))


@app.route("/admin/users/delete/<int:user_id>", methods=["POST"])
@admin_required
def admin_delete_user(user_id):
    conn = get_db()
    user = conn.execute("SELECT display_name FROM users WHERE id = ?", (user_id,)).fetchone()
    conn.execute("DELETE FROM users WHERE id = ?", (user_id,))
    conn.commit()
    conn.close()
    if user:
        flash(f"'{user['display_name']}' 계정이 삭제되었습니다.", "success")
    return redirect(url_for("admin_users"))


@app.route("/admin/users/reset-password/<int:user_id>", methods=["POST"])
@admin_required
def admin_reset_password(user_id):
    new_password = request.form.get("new_password", "").strip()
    if not new_password:
        flash("새 비밀번호를 입력해주세요.", "error")
        return redirect(url_for("admin_users"))
    pw_hash = generate_password_hash(new_password)
    conn = get_db()
    conn.execute("UPDATE users SET password_hash = ? WHERE id = ?", (pw_hash, user_id))
    conn.commit()
    conn.close()
    flash("비밀번호가 변경되었습니다.", "success")
    return redirect(url_for("admin_users"))


# 기존 /manage 경로 하위 호환 유지
@app.route("/manage")
@admin_required
def manage():
    return redirect(url_for("admin_items"))


@app.route("/history")
@login_required
def history():
    display_name = session.get("display_name", "")
    conn = get_db()
    rows = conn.execute("""
        SELECT r.id,
               i.name          AS item_name,
               i.type          AS item_type,
               r.borrower_name,
               r.rented_at,
               r.returned_at,
               r.due_date,
               r.cancelled_at
        FROM rentals r
        JOIN items i ON i.id = r.item_id
        WHERE r.borrower_name = ?
        ORDER BY r.id DESC
    """, (display_name,)).fetchall()
    conn.close()
    return render_template("history.html", rows=rows)


@app.route("/history/export")
@login_required
def history_export():
    import io
    from flask import send_file
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    display_name = session.get("display_name", "")
    conn = get_db()
    rows = conn.execute("""
        SELECT r.id, i.name AS item_name, i.type AS item_type,
               r.borrower_name, r.rented_at, r.due_date, r.returned_at, r.cancelled_at
        FROM rentals r JOIN items i ON i.id = r.item_id
        WHERE r.borrower_name = ?
        ORDER BY r.id DESC
    """, (display_name,)).fetchall()
    conn.close()

    TYPE_LABELS = {"drone": "드론", "laptop": "노트북", "camera": "카메라", "etc": "기타"}

    def fmt(val):
        if not val:
            return "—"
        try:
            from datetime import datetime as dt
            d = dt.fromisoformat(str(val)[:10])
            return f"{d.year % 100:02d}년 {d.month:02d}월 {d.day:02d}일"
        except Exception:
            return val

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "대여 이력"

    # 헤더 스타일
    header_fill = PatternFill("solid", fgColor="1E293B")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["No.", "물품명", "종류", "대여자", "대여일", "반납 예정일", "상태"]
    col_widths = [6, 16, 10, 12, 16, 16, 10]

    # 타이틀 행
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = "부서 공유 물품 대여 이력"
    title_cell.font = Font(bold=True, size=14, color="1E293B")
    title_cell.alignment = center
    ws.row_dimensions[1].height = 32

    # 헤더 행
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[2].height = 22

    # 데이터 행
    alt_fill = PatternFill("solid", fgColor="F8FAFC")
    for idx, r in enumerate(rows, 1):
        row_num = idx + 2
        status = "반납완료" if r["returned_at"] else "대여 중"
        values = [
            idx,
            r["item_name"],
            TYPE_LABELS.get(r["item_type"], "기타"),
            r["borrower_name"],
            fmt(r["rented_at"]),
            fmt(r["due_date"]),
            status,
        ]
        fill = alt_fill if idx % 2 == 0 else None
        status_color = "27AE60" if r["returned_at"] else "E67E22"

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.alignment = center
            cell.border = border
            if fill:
                cell.fill = fill
            if col == 7:  # 상태 컬럼 색상
                cell.font = Font(bold=True, color=status_color)
        ws.row_dimensions[row_num].height = 20

    # 출력일 표기
    from datetime import date
    ws.cell(row=len(rows) + 4, column=1,
            value=f"출력일: {date.today().strftime('%Y년 %m월 %d일')}").font = Font(color="94A3B8", size=9)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"대여이력_{date.today().strftime('%Y%m%d')}.xlsx"
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=filename)


@app.route("/qr")
def qr_page():
    base_url = request.host_url.rstrip("/")
    return render_template("qr.html", base_url=base_url)


@app.route("/qr/image")
def qr_image():
    import io, qrcode
    from flask import send_file
    base_url = request.host_url.rstrip("/")
    img = qrcode.make(base_url)
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    return send_file(buf, mimetype="image/png")


if __name__ == "__main__":
    app.run(debug=True)
